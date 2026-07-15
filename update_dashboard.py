import os
import yfinance as yf
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from datetime import datetime, date


#--------------------------
# Constants
#--------------------------

# File Settings
WORKBOOK_FOLDER = os.path.expanduser("~/OneDrive/Active")
WORKBOOK_NAME = "Financial Dashboard.xlsx"

# Worksheet Names
DASHBOARD_SHEET = "Dashboard"
MARKET_SHEET = "Market Data"
HISTORY_SHEET = "History"
ACTIVITY_SHEET = "Activity Log"
HOLDINGS_SHEET = "Holdings"
TSP_PRICE_SHEET = "TSP Prices"

# Dashboard Colors
GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
YELLOW = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
RED = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

# Dashboard Thresholds
GREEN_THRESHOLD = -0.01  # Within 1% of record high
YELLOW_THRESHOLD = -0.03  # Within 3% of record high

# Markets 
MARKETS = {
   "Dow": {
      "ticker": "^DJI",
      "row": 2
   },
   "S&P 500": {
     "ticker": "^GSPC",
       "row": 3
   },
   "Nasdaq": { 
   "ticker": "^IXIC",
   "row": 4
   }
}

# Dashboard Cells Mapping
DASHBOARD_CELLS = {
   "Dow": {
      "price": "B6",
      "status": "C6"
   },
   "S&P 500": {
      "price": "B7",
      "status": "C7"
   },
   "Nasdaq": {
      "price": "B8",
      "status": "C8"
   }
}

def open_workbook():
    file = os.path.join(WORKBOOK_FOLDER, WORKBOOK_NAME)
    if not os.path.exists(file):
        raise FileNotFoundError(f"Workbook not found at {file}. Please ensure the file exists.")
    wb = load_workbook(file)
    return wb, file

def get_market_sheet(wb):
    return wb[MARKET_SHEET]

def get_dashboard_sheet(wb):
    return wb[DASHBOARD_SHEET]

def get_history_sheet(wb):
    for name in wb.sheetnames:
        if name.strip().lower() == "history":
            return wb[name]
    return wb.create_sheet(HISTORY_SHEET)

def get_activity_sheet(wb):
    return wb[ACTIVITY_SHEET]

def get_holdings_sheet(wb):
    return wb[HOLDINGS_SHEET]

#--------------------------
# Update Market Data
#--------------------------

def update_dashboard(ws_dashboard, values):
    ws_dashboard["F2"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    for market, cells in DASHBOARD_CELLS.items():
        ws_dashboard[cells["price"]] = values[market]["price"]
        ws_dashboard[cells["status"]] = values[market]["pct_from_high"]
        ws_dashboard[cells["status"]].number_format = "0.00%"

        apply_color(
            ws_dashboard, 
            cells["status"],
            values[market]["pct_from_high"]
        )


def apply_color(ws, cell, pct):
    if pct >= GREEN_THRESHOLD:
        ws[cell].fill = GREEN
    elif pct >= YELLOW_THRESHOLD:
        ws[cell].fill = YELLOW
    else:
        ws[cell].fill = RED

def log_activity(ws_log, action, status = "Success"):
   ws_log.append([
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        action,
        status
    ])

def update_market_sheet(ws_market):
   values = {}

   for name, info in MARKETS.items():
      symbol = info["ticker"]
      row = info["row"]
      
      data = yf.Ticker(symbol).history(period="max")

      close = data['Close']
      high = data['High']

      current = round(close.iloc[-1], 2)
      record_close = round(close.max(), 2)
      record_close_date = close.idxmax().date()
      record_high = high.max()
      pct_from_high = (current / record_high) - 1


      # Prices
      ws_market[f"B{row}"] = current
      ws_market[f"C{row}"] = record_close

      # Date 
      ws_market[f"D{row}"] = record_close_date
      ws_market[f"D{row}"].number_format = "mmm d, yyyy"

      # Percent from high (formatted properly)
      ws_market[f"E{row}"] = pct_from_high
      ws_market[f"E{row}"].number_format = "0.00%"

      # Save for history log
      values[name] = {
          "price": current,
          "pct_from_high": pct_from_high
       }
   return values

# --------------------------
# Append to History Sheet
#--------------------------

def update_history(ws_history, values):

    row = [
        date.today(),
        values["Dow"]["price"],
        values["S&P 500"]["price"],
        values["Nasdaq"]["price"]
    ]

    ws_history.append(row)

def update_holdings (ws_holdings):
   for row in range(2, ws_holdings.max_row + 1):

       symbol = ws_holdings[f"B{row}"].value
       shares = ws_holdings[f"D{row}"].value
       source = ws_holdings[f"G{row}"].value 

       source = str(ws_holdings[f"G{row}"].value).strip().lower()

       if source == "yahoo":
            try:
                data = yf.Ticker(symbol).history(period="1d")
            
                if data.empty:
                    print (f"No data for {symbol}")
                    continue

                price = round(data['Close'].iloc[-1], 2)

                ws_holdings[f"E{row}"] = price

                if shares:
                    ws_holdings[f"F{row}"] = round(price * shares, 2)
 
            except Exception as e:
                print(f"{symbol}: {e}")


def save_workbook(wb, file):
    wb.save(file)

def main():
    try:
         wb, file = open_workbook()

         ws_market = get_market_sheet(wb)
         ws_dashboard = get_dashboard_sheet(wb)
         ws_history = get_history_sheet(wb)
         ws_log = get_activity_sheet(wb)
         ws_holdings = get_holdings_sheet(wb)

         log_activity(ws_log, "Starting update")

         values = update_market_sheet(ws_market)
         log_activity(ws_log, "Market data updated")

         update_holdings(ws_holdings)
         log_activity(ws_log, "Holdings updated")

         update_history(ws_history, values)
         log_activity(ws_log, "History updated")

         update_dashboard(ws_dashboard, values)
         log_activity(ws_log, "Dashboard updated")

         log_activity(ws_log, "Workbook saved")
         save_workbook(wb, file)

         print("Dashboard updated successfully.")
    except Exception as e:
         print(f"Error: {e}")

if __name__ == "__main__":
    main()
